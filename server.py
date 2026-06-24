"""
L&T EPS ADAS Partner Evaluation Dashboard — Local Server
Run: python server.py
Open: http://localhost:5000
"""

import csv
import io
import json
import os
import re
import sys
import uuid
import shutil
from datetime import date, datetime
from pathlib import Path

try:
    from flask import Flask, jsonify, request, send_from_directory, abort, Response
except ImportError:
    print("Flask not found. Installing...")
    os.system(f"{sys.executable} -m pip install flask")
    from flask import Flask, jsonify, request, send_from_directory, abort

from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).parent.resolve()
DATA_DIR = BASE_DIR / "data"
STATIC_DIR = BASE_DIR / "static"
PARTNERS_FILE = DATA_DIR / "partners.json"
SCHEMA_FILE = BASE_DIR / "schema.json"
EXPORTS_LOG_FILE = DATA_DIR / "questionnaire_exports.json"

DATA_DIR.mkdir(exist_ok=True)

app = Flask(__name__, static_folder=str(STATIC_DIR))


# Every client-side api() call expects JSON back, even on failure -- without
# this, an unhandled exception or a plain abort(404) returns Flask/Werkzeug's
# default HTML error page, and `res.json()` on the client blows up with a
# confusing "Unexpected token '<'" instead of the real error. Found via a
# real crash on a genuine customer-filled import (see api_import_questionnaire_preview).
@app.errorhandler(Exception)
def handle_any_error(e):
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return jsonify({"error": e.description or str(e)}), e.code
    app.logger.exception("Unhandled exception")
    return jsonify({"error": "Internal server error: " + str(e)}), 500


# ── helpers ──────────────────────────────────────────────────────────────────

def load_partners():
    if not PARTNERS_FILE.exists():
        return {"version": 1, "partners": []}
    with open(PARTNERS_FILE, encoding="utf-8") as f:
        return json.load(f)


def save_partners(data):
    # Atomic write via temp file to prevent corruption on crash
    tmp = PARTNERS_FILE.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    shutil.move(str(tmp), str(PARTNERS_FILE))


def load_schema():
    with open(SCHEMA_FILE, encoding="utf-8") as f:
        return json.load(f)


def save_schema(data):
    # Bumped on every write so an exported questionnaire can detect "the
    # schema changed since this file was generated" with a plain != check
    # (see plan.md Step 5.2) -- every save_schema() call site already
    # represents a real schema mutation, so bumping unconditionally here
    # is correct.
    data["schemaVersion"] = data.get("schemaVersion", 0) + 1
    # Atomic write via temp file to prevent corruption on crash
    tmp = SCHEMA_FILE.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    shutil.move(str(tmp), str(SCHEMA_FILE))


# Append-only local log of every questionnaire export (plan.md Step 5.4).
# Each entry's per-sheet qIds breakdown is what lets a later session answer
# "was this question even in the file I sent" -- distinguishing a customer
# leaving something blank on purpose from a question that didn't exist yet
# in their copy.
def load_export_log():
    if not EXPORTS_LOG_FILE.exists():
        return []
    with open(EXPORTS_LOG_FILE, encoding="utf-8") as f:
        return json.load(f)


def append_export_log_entry(entry):
    log = load_export_log()
    log.append(entry)
    tmp = EXPORTS_LOG_FILE.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(log, f, ensure_ascii=False, indent=2)
    shutil.move(str(tmp), str(EXPORTS_LOG_FILE))


# ── static files ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")


@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory(STATIC_DIR, filename)


# ── schema ───────────────────────────────────────────────────────────────────

@app.route("/api/schema")
def api_schema():
    return jsonify(load_schema())


# Promotes a partner-local draft question into the global question list.
# Always appends (never inserts mid-array) so existing questions' dotted
# numbering (schema.domains -> detailSections -> detailQuestions order)
# never shifts. sourceRef is null since this didn't come from a source
# spreadsheet, same as the other manually-authored questions already in
# schema.json.
@app.route("/api/schema/questions", methods=["POST"])
def api_publish_question():
    body = request.get_json(force=True)
    section_id = (body.get("sectionId") or "").strip()
    text = (body.get("text") or "").strip()
    priority = body.get("priority") or "medium"
    if not text:
        return jsonify({"error": "text is required"}), 400
    if priority not in ("high", "medium", "low"):
        priority = "medium"

    schema = load_schema()
    if not any(s["id"] == section_id for s in schema.get("detailSections", [])):
        return jsonify({"error": "unknown sectionId"}), 400

    next_id = max([q["id"] for q in schema.get("detailQuestions", [])], default=0) + 1
    new_question = {
        "id": next_id,
        "sectionId": section_id,
        "text": text,
        "sourceRef": None,
        "priority": priority,
    }
    schema["detailQuestions"].append(new_question)
    save_schema(schema)
    return jsonify(new_question), 201


# Edits an existing question's priority -- the only field this exposes for
# now (text edits would need an "edit question" UI that doesn't exist yet;
# removing a question already has its own publish->usage->delete flow
# above). Schema-level, so it takes effect for every partner immediately,
# same as publish/remove/add-function.
@app.route("/api/schema/questions/<int:q_id>", methods=["PUT"])
def api_update_question(q_id):
    body = request.get_json(force=True)
    priority = body.get("priority")
    if priority not in ("high", "medium", "low"):
        return jsonify({"error": "priority must be one of high/medium/low"}), 400

    schema = load_schema()
    question = next((q for q in schema.get("detailQuestions", []) if q["id"] == q_id), None)
    if not question:
        abort(404)

    question["priority"] = priority
    save_schema(schema)
    return jsonify(question)


def _answer_has_data(a):
    return bool(a.get("status") or (a.get("remarks") or "").strip()
                or (a.get("partnerResponse") or "").strip())


# Read-only impact check before a destructive removal — lets the client
# show exactly which partners/products already answered this question
# before the user commits to deleting it everywhere.
@app.route("/api/schema/questions/<int:q_id>/usage", methods=["GET"])
def api_question_usage(q_id):
    data = load_partners()
    usages = []
    for p in data["partners"]:
        for a in p.get("generalAnswers", []):
            if a.get("qId") == q_id and _answer_has_data(a):
                usages.append({"partnerName": p["name"], "scope": "general"})
        for prod in p.get("products", []):
            for a in prod.get("answers", []):
                if a.get("qId") == q_id and _answer_has_data(a):
                    usages.append({
                        "partnerName": p["name"],
                        "scope": "product",
                        "productName": prod.get("name", "Product"),
                    })
    return jsonify({"usages": usages})


# Removes a published question everywhere — from schema.detailQuestions
# and from every partner's saved answer to it (general or per-product),
# not just the partner currently open client-side. The client is expected
# to have already shown the impact (via the usage endpoint above) and
# confirmed before calling this.
@app.route("/api/schema/questions/<int:q_id>", methods=["DELETE"])
def api_remove_question(q_id):
    schema = load_schema()
    before = len(schema.get("detailQuestions", []))
    schema["detailQuestions"] = [q for q in schema.get("detailQuestions", []) if q["id"] != q_id]
    if len(schema["detailQuestions"]) == before:
        abort(404)
    save_schema(schema)

    data = load_partners()
    swept = 0
    for p in data["partners"]:
        before_g = len(p.get("generalAnswers", []))
        p["generalAnswers"] = [a for a in p.get("generalAnswers", []) if a.get("qId") != q_id]
        swept += before_g - len(p["generalAnswers"])
        for prod in p.get("products", []):
            before_p = len(prod.get("answers", []))
            prod["answers"] = [a for a in prod.get("answers", []) if a.get("qId") != q_id]
            swept += before_p - len(prod["answers"])
    save_partners(data)
    return jsonify({"ok": True, "answersRemoved": swept})


# Adds a brand-new function (e.g. "Surround View") to the shared schema,
# plus a matching empty product-scoped section -- mirrors "AEB Function
# Detail" exactly (category "product", scope {type: "function", keys:
# [key]}), so it picks up every existing rendering/scope-matching code
# path with no extra propagation logic, same reasoning as publishing a
# question. Starts with zero questions on purpose -- partners add their
# own via the existing draft -> Publish flow once the function is real.
@app.route("/api/schema/functions", methods=["POST"])
def api_add_function():
    body = request.get_json(force=True)
    label = (body.get("label") or "").strip()
    if not label:
        return jsonify({"error": "label is required"}), 400

    key = re.sub(r"[^A-Z0-9]+", "_", label.upper()).strip("_")
    if not key:
        return jsonify({"error": "could not derive a valid key from that name"}), 400

    schema = load_schema()
    if any(f["key"] == key for f in schema.get("productFunctions", [])):
        return jsonify({"error": f'a function named "{label}" already exists'}), 400

    section_id = "prod-func-" + key.lower().replace("_", "-")
    if any(s["id"] == section_id for s in schema.get("detailSections", [])):
        return jsonify({"error": "a section with that id already exists"}), 400

    new_section = {
        "id": section_id,
        "category": "product",
        "label": f"{label} Function Detail",
        "domain": "perception-function",
        "scope": {"type": "function", "keys": [key]},
    }
    schema["productFunctions"].append({"key": key, "label": label})
    schema["detailSections"].append(new_section)
    save_schema(schema)
    return jsonify({"key": key, "label": label, "section": new_section}), 201


# ── partners list ─────────────────────────────────────────────────────────────

@app.route("/api/partners", methods=["GET"])
def api_partners_list():
    data = load_partners()
    summary = []
    for p in data["partners"]:
        summary.append({
            "id": p["id"],
            "name": p["name"],
            "shortName": p.get("shortName", p["name"][:2].upper()),
            "stage": p.get("stage", "Under Evaluation"),
            "evaluationDate": p.get("evaluationDate"),
            "evaluator": p.get("evaluator", ""),
            "decisionStatus": p.get("decision", {}).get("status", ""),
            "productCount": len(p.get("products", [])),
        })
    return jsonify(summary)


@app.route("/api/partners/<partner_id>", methods=["GET"])
def api_partner_detail(partner_id):
    data = load_partners()
    partner = next((p for p in data["partners"] if p["id"] == partner_id), None)
    if not partner:
        abort(404)
    return jsonify(partner)


# ── create partner ────────────────────────────────────────────────────────────

@app.route("/api/partners", methods=["POST"])
def api_create_partner():
    body = request.get_json(force=True)
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400

    new_partner = {
        "id": str(uuid.uuid4()),
        "name": name,
        "shortName": body.get("shortName", name[:2].upper()),
        "addedDate": date.today().isoformat(),
        "lastModified": datetime.now().isoformat(),
        "evaluator": body.get("evaluator", ""),
        "productVersion": body.get("productVersion", ""),
        "evaluationDate": body.get("evaluationDate", ""),
        "stage": body.get("stage", "Under Evaluation"),
        "businessModel": [],
        "products": [],
        "generalAnswers": [],
        "draftQuestions": [],
        "sectionGrades": [],
        "domainGrades": [],
        "patents": [],
        "hardDisqualifiers": [],
        "decision": {
            "rationale": "",
            "reviewer": "",
            "date": "",
            "status": "",
        },
        "notes": body.get("notes", ""),
    }

    data = load_partners()
    data["partners"].append(new_partner)
    save_partners(data)
    return jsonify({"id": new_partner["id"]}), 201


# ── update partner ────────────────────────────────────────────────────────────

@app.route("/api/partners/<partner_id>", methods=["PUT"])
def api_update_partner(partner_id):
    data = load_partners()
    idx = next((i for i, p in enumerate(data["partners"]) if p["id"] == partner_id), None)
    if idx is None:
        abort(404)

    # Optimistic concurrency (plan.md Step 5.3): if the client's loaded
    # lastModified doesn't match what's currently on disk, something else
    # changed this partner since it was loaded (another tab, or a
    # questionnaire import) -- reject rather than silently overwrite.
    # A partner with no lastModified yet (pre-Step-5 data) is treated as
    # compatible with anything, so existing partners don't break on their
    # first save after this field was introduced.
    current_last_modified = data["partners"][idx].get("lastModified")
    body = request.get_json(force=True)
    submitted_last_modified = body.get("lastModified")
    if current_last_modified and submitted_last_modified != current_last_modified:
        return jsonify({
            "error": "This partner changed since you loaded it. Reload before saving.",
        }), 409

    # Preserve id and addedDate
    body["id"] = partner_id
    body.setdefault("addedDate", data["partners"][idx].get("addedDate", date.today().isoformat()))
    body["lastModified"] = datetime.now().isoformat()

    data["partners"][idx] = body
    save_partners(data)
    return jsonify({"ok": True, "lastModified": body["lastModified"]})


# ── delete partner ────────────────────────────────────────────────────────────

@app.route("/api/partners/<partner_id>", methods=["DELETE"])
def api_delete_partner(partner_id):
    data = load_partners()
    before = len(data["partners"])
    data["partners"] = [p for p in data["partners"] if p["id"] != partner_id]
    if len(data["partners"]) == before:
        abort(404)
    save_partners(data)
    return jsonify({"ok": True})


# ── Questionnaire export/import (plan.md Step 5) ────────────────────────────
# Python ports of app.js's sectionScopeMatches/sectionHasAnswerData/
# questionLabel -- the export and the in-app Product tab must agree on
# exactly which questions are "relevant" to a product and how they're
# numbered, so these mirror the JS logic line for line rather than
# reinventing it server-side.

def scope_field(scope_type):
    return {"sensor": "sensors", "function": "functions", "soc": "socs"}.get(scope_type)


def section_scope_matches(section, prod):
    scope = section.get("scope")
    if not scope or not scope.get("keys"):
        return True
    field = scope_field(scope["type"])
    prod_field = prod.get(field) or {}
    return any(prod_field.get(key) for key in scope["keys"])


def section_has_answer_data(section, prod, schema):
    questions = [q for q in schema["detailQuestions"] if q["sectionId"] == section["id"]]
    answers = prod.get("answers", [])
    for q in questions:
        a = next((x for x in answers if x.get("qId") == q["id"]), None)
        if a and (a.get("status") or (a.get("remarks") or "").strip() or (a.get("partnerResponse") or "").strip()):
            return True
    return False


def sections_for_product(prod, schema):
    sections = [s for s in schema["detailSections"] if s["category"] == "product"]
    return [s for s in sections if section_scope_matches(s, prod) or section_has_answer_data(s, prod, schema)]


def question_label(q, schema):
    section = next((s for s in schema["detailSections"] if s["id"] == q["sectionId"]), None)
    if not section:
        return str(q["id"])
    questions_in_section = [qq["id"] for qq in schema["detailQuestions"] if qq["sectionId"] == q["sectionId"]]
    question_num = questions_in_section.index(q["id"]) + 1
    domains = schema["domains"]
    if section["category"] == "general":
        cat_domains = [d["id"] for d in domains if d["category"] == "general"]
        cat_num = (cat_domains.index(section["domain"]) + 1) if section["domain"] in cat_domains else 0
        if cat_num <= 0 or question_num <= 0:
            return str(q["id"])
        return f"{cat_num}.{question_num}"
    cat_domains = [d["id"] for d in domains if d["category"] == "product"]
    cat_num = (cat_domains.index(section["domain"]) + 1) if section["domain"] in cat_domains else 0
    sections_in_domain = [s["id"] for s in schema["detailSections"] if s["domain"] == section["domain"]]
    section_num = (sections_in_domain.index(section["id"]) + 1) if section["id"] in sections_in_domain else 0
    if cat_num <= 0 or section_num <= 0 or question_num <= 0:
        return str(q["id"])
    return f"{cat_num}.{section_num}.{question_num}"


def ordered_questions_for_sections(sections, schema):
    # Same order questionLabel()'s numbers ascend in: schema.domains array
    # order first (mirrors domainGroupedSectionsHtml's grouping in the UI),
    # then each domain's own sections in their relative detailSections
    # order, then question order within section. Sections are NOT
    # contiguous in detailSections by domain any more (Step 4 reassigned
    # domain/category per-section without reordering the array), so
    # grouping by domain explicitly here is required -- a flat scan of
    # detailSections array order would interleave domains out of numeric
    # order (confirmed by a real test export: domain 5 rows appeared
    # before domain 4 and domain 1 rows).
    section_ids = {s["id"] for s in sections}
    by_domain = {}
    for s in schema["detailSections"]:
        if s["id"] in section_ids:
            by_domain.setdefault(s["domain"], []).append(s["id"])
    out = []
    for d in schema["domains"]:
        for sid in by_domain.get(d["id"], []):
            out.extend(q for q in schema["detailQuestions"] if q["sectionId"] == sid)
    return out


SHEET_NAME_BAD_CHARS = ':\\/?*[]'


def safe_sheet_name(name):
    cleaned = "".join(c for c in (name or "Product") if c not in SHEET_NAME_BAD_CHARS)
    return (cleaned.strip() or "Product")[:31]


# Sheet-level identity is stored in fixed hidden cells (column G label,
# column H value) rather than per-row, since it's the same for every row on
# a given sheet -- partnerId/schemaVersion/exportId are workbook-wide
# (same on every sheet), productId is per-sheet (None on the General sheet).
def _write_metadata(ws, partner_id, product_id, schema_version, export_id):
    ws["G1"] = "partnerId"
    ws["H1"] = partner_id
    ws["G2"] = "productId"
    ws["H2"] = product_id or ""
    ws["G3"] = "schemaVersion"
    ws["H3"] = schema_version
    ws["G4"] = "exportId"
    ws["H4"] = export_id
    ws.column_dimensions["E"].hidden = True
    ws.column_dimensions["G"].hidden = True
    ws.column_dimensions["H"].hidden = True


def _read_metadata(ws):
    return {
        "partnerId": ws["H1"].value,
        "productId": ws["H2"].value or None,
        "schemaVersion": ws["H3"].value,
        "exportId": ws["H4"].value,
    }


def _build_sheet(wb, title, questions, schema, partner_id, product_id, schema_version, export_id):
    ws = wb.create_sheet(title=safe_sheet_name(title))
    headers = ["#", "Question", "Partner Response", "Status", "qId"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    status_labels = [s["label"] for s in schema.get("answerStatuses", [])]
    dv = None
    if status_labels:
        formula = '"' + ",".join(status_labels) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)

    for i, q in enumerate(questions, start=2):
        ws.cell(row=i, column=1, value=question_label(q, schema))
        ws.cell(row=i, column=2, value=q["text"])
        resp_cell = ws.cell(row=i, column=3, value="")
        resp_cell.protection = Protection(locked=False)
        status_cell = ws.cell(row=i, column=4, value="")
        status_cell.protection = Protection(locked=False)
        if dv:
            dv.add(status_cell)
        ws.cell(row=i, column=5, value=q["id"])

    _write_metadata(ws, partner_id, product_id, schema_version, export_id)

    ws.protection.sheet = True
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.insertColumns = False
    ws.protection.deleteColumns = False
    ws.protection.sort = False
    ws.protection.formatCells = False
    return [q["id"] for q in questions]


@app.route("/api/partners/<partner_id>/export-questionnaire", methods=["GET"])
def api_export_questionnaire(partner_id):
    data = load_partners()
    partner = next((p for p in data["partners"] if p["id"] == partner_id), None)
    if not partner:
        abort(404)
    schema = load_schema()
    schema_version = schema.get("schemaVersion", 0)
    export_id = str(uuid.uuid4())

    wb = Workbook()
    wb.remove(wb.active)
    sheets_log = []

    general_sections = [s for s in schema["detailSections"] if s["category"] == "general"]
    general_questions = ordered_questions_for_sections(general_sections, schema)
    qids = _build_sheet(wb, "General", general_questions, schema, partner_id, None, schema_version, export_id)
    sheets_log.append({"productId": None, "qIds": qids})

    for prod in partner.get("products", []):
        sections = sections_for_product(prod, schema)
        questions = ordered_questions_for_sections(sections, schema)
        qids = _build_sheet(wb, prod.get("name") or "Product", questions, schema,
                             partner_id, prod["id"], schema_version, export_id)
        sheets_log.append({"productId": prod["id"], "qIds": qids})

    append_export_log_entry({
        "exportId": export_id,
        "partnerId": partner_id,
        "schemaVersion": schema_version,
        "exportedAt": datetime.now().isoformat(),
        "sheets": sheets_log,
    })

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"questionnaire_{partner['name']}_{date.today().isoformat()}.xlsx".replace(" ", "_")
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def cell_str(value):
    # Excel cells can hold a typed value (int/float/bool/datetime) even in
    # a column meant for free text -- nothing stops a customer from typing
    # a bare number into Partner Response. Found via a real import crash:
    # 'int' object has no attribute 'strip'.
    if value is None:
        return ""
    return str(value).strip()


def status_key_from_label(label, schema):
    if not label:
        return None
    label = str(label).strip()
    for s in schema.get("answerStatuses", []):
        if s["label"] == label:
            return s["key"]
    return None


# Preview-only: parses and validates, writes nothing. The client shows this
# diff and only calls /commit (below) with the exact `changes` list the
# user confirmed -- commit never re-parses the file, so a file swapped
# between the two calls can't sneak in unreviewed changes.
@app.route("/api/partners/<partner_id>/import-questionnaire/preview", methods=["POST"])
def api_import_questionnaire_preview(partner_id):
    data = load_partners()
    partner = next((p for p in data["partners"] if p["id"] == partner_id), None)
    if not partner:
        abort(404)

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        wb = load_workbook(BytesIO(file.read()))
    except Exception:
        return jsonify({"error": "Could not read this file. Make sure it's an .xlsx exported from this tool."}), 400

    schema = load_schema()
    current_schema_version = schema.get("schemaVersion", 0)
    questions_by_id = {q["id"]: q for q in schema["detailQuestions"]}
    product_ids = {p["id"] for p in partner.get("products", [])}

    # Hard reject before anything else is computed -- every sheet's
    # partnerId must agree with the partner being imported into. A file
    # with no partnerId on *any* sheet isn't a mismatch, it's not a
    # questionnaire from this tool at all (found via testing: a random
    # unrelated .xlsx was silently accepted as "0 changes," which looks
    # exactly like a customer who genuinely left everything blank --
    # misleading, since those are very different situations).
    any_partner_id = False
    for name in wb.sheetnames:
        meta = _read_metadata(wb[name])
        if meta["partnerId"]:
            any_partner_id = True
            if meta["partnerId"] != partner_id:
                return jsonify({"error": "This file was exported for a different partner."}), 400
    if not any_partner_id:
        return jsonify({"error": "This doesn't look like a questionnaire exported from this tool."}), 400

    changes = []
    flagged = []
    schema_version_at_export = None
    for name in wb.sheetnames:
        ws = wb[name]
        meta = _read_metadata(ws)
        schema_version_at_export = meta["schemaVersion"]
        product_id = meta["productId"]

        if product_id and product_id not in product_ids:
            flagged.append({"sheet": name, "reason": "Product no longer exists on this partner", "skippedWholeSheet": True})
            continue

        for row in ws.iter_rows(min_row=2, values_only=False):
            raw_q_id = row[4].value  # column E
            if raw_q_id is None:
                continue
            try:
                q_id = int(raw_q_id)
            except (TypeError, ValueError):
                flagged.append({"sheet": name, "reason": f"Row has an unreadable qId ({raw_q_id!r}) -- skipped"})
                continue
            question = questions_by_id.get(q_id)
            if not question:
                flagged.append({"sheet": name, "qId": q_id, "reason": "Question no longer exists in the current schema"})
                continue

            file_text = cell_str(row[1].value)
            if file_text != question["text"].strip():
                flagged.append({"sheet": name, "qId": q_id, "reason": "Question text doesn't match the current schema (may have changed since export, or the file's structure was altered)"})
                continue

            partner_response = cell_str(row[2].value)
            status_label = cell_str(row[3].value)
            status_key = status_key_from_label(status_label, schema)
            if status_label and not status_key:
                flagged.append({"sheet": name, "qId": q_id, "reason": f'Status value "{status_label}" doesn\'t match any current status option -- ignored, response (if any) still applied'})

            if not partner_response and not status_key:
                continue  # nothing to apply for this row

            changes.append({
                "scope": "product" if product_id else "general",
                "productId": product_id,
                "qId": q_id,
                "questionText": question["text"],
                "partnerResponse": partner_response or None,
                "status": status_key,
            })

    return jsonify({
        "ok": True,
        "partnerId": partner_id,
        "schemaVersionAtExport": schema_version_at_export,
        "currentSchemaVersion": current_schema_version,
        "schemaVersionMismatch": schema_version_at_export is not None and schema_version_at_export != current_schema_version,
        "changes": changes,
        "flagged": flagged,
    })


# Applies exactly the `changes` list the client got back from /preview and
# the user confirmed -- never re-parses a file. PATCH semantics: only the
# fields present and non-empty in each change are written, so a partially
# filled re-send can never blank out an answer that already exists.
@app.route("/api/partners/<partner_id>/import-questionnaire/commit", methods=["POST"])
def api_import_questionnaire_commit(partner_id):
    data = load_partners()
    idx = next((i for i, p in enumerate(data["partners"]) if p["id"] == partner_id), None)
    if idx is None:
        abort(404)
    partner = data["partners"][idx]

    body = request.get_json(force=True)
    changes = body.get("changes") or []
    updated = 0

    for ch in changes:
        q_id = ch["qId"]
        if ch["scope"] == "general":
            answers = partner.setdefault("generalAnswers", [])
            ans = next((a for a in answers if a.get("qId") == q_id), None)
            if not ans:
                ans = {"qId": q_id, "status": "", "remarks": "", "partnerResponse": ""}
                answers.append(ans)
        else:
            prod = next((pr for pr in partner.get("products", []) if pr["id"] == ch["productId"]), None)
            if not prod:
                continue  # product was deleted between preview and commit
            answers = prod.setdefault("answers", [])
            ans = next((a for a in answers if a.get("qId") == q_id), None)
            if not ans:
                ans = {"qId": q_id, "status": "", "remarks": "", "partnerResponse": ""}
                answers.append(ans)

        if ch.get("partnerResponse"):
            ans["partnerResponse"] = ch["partnerResponse"]
        if ch.get("status"):
            ans["status"] = ch["status"]
        updated += 1

    partner["lastModified"] = datetime.now().isoformat()
    save_partners(data)
    return jsonify({"ok": True, "updated": updated, "lastModified": partner["lastModified"]})


# ── CSV export ───────────────────────────────────────────────────────────────

@app.route("/api/export/csv")
def api_export_csv():
    data = load_partners()

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header row
    writer.writerow(
        ["Partner", "Stage", "Evaluator", "Eval Date", "Decision Status", "Reviewer", "Decision Date", "Rationale"]
    )

    for p in data["partners"]:
        d = p.get("decision") or {}
        writer.writerow(
            [p["name"], p.get("stage", ""), p.get("evaluator", ""), p.get("evaluationDate", "")]
            + [
                d.get("status", ""),
                d.get("reviewer", ""),
                d.get("date", ""),
                d.get("rationale", ""),
            ]
        )

    filename = f"adas_partner_evaluation_{date.today().isoformat()}.csv"
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import webbrowser
    port = 5000
    print(f"\n  L&T EPS ADAS Partner Evaluation Dashboard")
    print(f"  Open: http://localhost:{port}")
    print(f"  Press Ctrl+C to stop\n")
    webbrowser.open(f"http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
